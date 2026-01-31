#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Location: ./tests/manual/generate_test_plan.py
Copyright 2025
SPDX-License-Identifier: Apache-2.0
Authors: Mihai Criveti

MCP Gateway v0.9.0 - Test Plan Generator from YAML

Generates Excel test plan from YAML test definition files.
Much cleaner and more maintainable approach.

Usage:
    python3 generate_test_plan.py
"""

import sys
import yaml
from pathlib import Path


def generate_excel_from_yaml():
    """Generate Excel file from YAML test definitions."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Install: pip install openpyxl pyyaml")
        return False

    print("📊 GENERATING EXCEL FROM YAML TEST FILES")
    print("=" * 60)
    print("📁 Reading from testcases/ directory")

    # Find YAML files in testcases directory
    testcases_dir = Path("testcases")
    if not testcases_dir.exists():
        print("❌ testcases/ directory not found")
        return False

    yaml_files = list(testcases_dir.glob("*.yaml"))
    yaml_files = sorted(yaml_files)

    if not yaml_files:
        print("❌ No YAML test files found")
        return False

    print(f"📄 Found {len(yaml_files)} YAML files:")
    for yf in yaml_files:
        print(f"   📄 {yf.name}")

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Styles
    styles = {
        "title": Font(size=16, bold=True, color="1F4E79"),
        "header_fill": PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        "header_font": Font(color="FFFFFF", bold=True),
        "critical_fill": PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid"),
        "critical_font": Font(color="FFFFFF", bold=True),
    }

    # Process each YAML file
    for yaml_file in yaml_files:
        try:
            with open(yaml_file, "r") as f:
                yaml_data = yaml.safe_load(f)

            worksheet_name = yaml_data.get("worksheet_name", yaml_file.stem)
            headers = yaml_data.get("headers", [])
            tests = yaml_data.get("tests", [])

            print(f"\n   📄 {yaml_file.name} → {worksheet_name}")
            print(f"      📊 {len(tests)} tests")

            # Create worksheet
            sheet = wb.create_sheet(worksheet_name)

            # Add headers
            for i, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=i, value=header)
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]

            # Add test data
            for row_idx, test in enumerate(tests, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = get_yaml_value(test, header)
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                    # Apply formatting
                    if header.lower() == "priority" and value == "CRITICAL":
                        cell.fill = styles["critical_fill"]
                        cell.font = styles["critical_font"]
                    elif header.lower() == "status":
                        cell.value = "☐"

            # Auto-size columns
            for col in range(1, len(headers) + 1):
                max_len = 0
                for row in range(1, min(len(tests) + 2, 20)):
                    val = sheet.cell(row=row, column=col).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                width = min(max(max_len + 2, 10), 60)
                sheet.column_dimensions[get_column_letter(col)].width = width

            print("      ✅ Created")

        except Exception as e:
            print(f"      ❌ Failed: {e}")

    # Save file
    output_path = Path("test-plan.xlsx")

    try:
        print("\n💾 Saving Excel file...")
        wb.save(output_path)
        wb.close()  # CRITICAL: Close properly

        print(f"✅ File saved: {output_path}")

        # Verify
        test_wb = openpyxl.load_workbook(output_path)
        print(f"✅ Verified: {len(test_wb.worksheets)} worksheets")
        test_wb.close()

        print("\n🎊 SUCCESS! Excel generated from YAML files!")
        return True

    except Exception as e:
        print(f"❌ Save failed: {e}")
        return False


def get_yaml_value(test, header):
    """Get value from YAML test data for Excel header."""

    mappings = {
        "Test ID": "test_id",
        "Priority": "priority",
        "Component": "component",
        "Description": "description",
        "Detailed Steps": "steps",
        "Steps": "steps",
        "Expected Result": "expected",
        "Expected": "expected",
        "Endpoint": "endpoint",
        "Method": "method",
        "cURL Command": "curl_command",
        "Request Body": "request_body",
        "Expected Status": "expected_status",
        "Expected Response": "expected_response",
        "Attack Type": "attack_type",
        "Target": "target",
        "Risk Level": "risk_level",
        "Attack Steps": "attack_steps",
        "Expected Defense": "expected_defense",
    }

    yaml_key = mappings.get(header, header.lower().replace(" ", "_"))
    value = test.get(yaml_key, "")

    # Handle special cases
    if header in ["SQLite", "PostgreSQL"]:
        return "✓" if test.get(f"{header.lower()}_support", True) else "❌"
    elif header in ["Actual Output", "Actual Status", "Actual Response", "Tester", "Date", "Comments"]:
        return ""  # Empty for tester to fill
    elif header == "Status":
        return "☐"

    return str(value) if value else ""


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--help":
        print("📊 Test Plan Generator from YAML")
        print("Usage:")
        print("  python3 generate_test_plan.py    # Generate Excel from YAML")
        print("  python3 generate_test_plan.py --help  # This help")
        print("\nEdit YAML files to update tests, then regenerate Excel.")
    else:
        try:
            success = generate_excel_from_yaml()
            if not success:
                sys.exit(1)
        except Exception as e:
            print(f"❌ Error: {e}")
            sys.exit(1)
